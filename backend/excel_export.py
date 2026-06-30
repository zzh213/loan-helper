"""生成贷款方案 Excel 报告。使用 openpyxl,多工作表分类展示。"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import EnterpriseProfile, RecommendResponse

CREDIT_CN = {"excellent": "优秀", "good": "良好", "fair": "一般", "poor": "较差"}
PURPOSE_CN = {
    "working_capital": "流动资金周转", "equipment": "设备采购", "expansion": "扩大经营",
    "inventory": "备货采购", "rd": "研发投入", "other": "其他",
}
IMPACT_CN = {"positive": "正向", "negative": "负向", "neutral": "中性"}

HEADER_FILL = PatternFill("solid", fgColor="1E40AF")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1E40AF")
LABEL_FILL = PatternFill("solid", fgColor="F3F4F6")
LABEL_FONT = Font(bold=True, color="4B5563")
BEST_FILL = PatternFill("solid", fgColor="FEF3C7")
THIN = Side(style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_excel(profile: EnterpriseProfile, result: RecommendResponse) -> bytes:
    wb = Workbook()

    # ---------- 概览 ----------
    ws = wb.active
    ws.title = "方案概览"
    ws["A1"] = "中小微企业贷款方案报告"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    name = profile.company_name or "企业"
    ws["A2"] = f"{name} · 生成时间 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(color="6B7280", size=9)
    ws.merge_cells("A2:D2")

    ws["A4"] = "方案摘要"
    ws["A4"].font = LABEL_FONT
    ws["A5"] = result.summary
    ws["A5"].alignment = WRAP
    ws.merge_cells("A5:D8")

    r = result.risk
    ws["A10"] = "综合风控评分"
    ws["A10"].font = LABEL_FONT
    ws["B10"] = f"{r.score} / 100"
    ws["C10"] = "风险等级"
    ws["C10"].font = LABEL_FONT
    ws["D10"] = f"{r.grade}（{r.grade_label}）"
    if result.plans:
        best = result.plans[0]
        ws["A11"] = "最优推荐"
        ws["A11"].font = LABEL_FONT
        ws["B11"] = best.product_name
        ws["C11"] = "预估额度/利率"
        ws["C11"].font = LABEL_FONT
        ws["D11"] = f"{best.estimated_amount}万 / {best.annual_rate_min}%-{best.annual_rate_max}%"
    _set_widths(ws, [18, 26, 18, 30])

    # ---------- 企业画像 ----------
    ws2 = wb.create_sheet("企业画像")
    rows = [
        ("企业名称", profile.company_name or "-", "所属行业", profile.industry),
        ("经营年限", f"{profile.years_in_business} 年", "年营业额", f"{profile.annual_revenue} 万元"),
        ("注册资本", f"{profile.registered_capital} 万元", "员工人数", f"{profile.employees} 人"),
        ("征信状况", CREDIT_CN.get(profile.credit_level.value, "-"),
         "当前逾期", "是" if profile.has_overdue else "否"),
        ("抵押物", f"{profile.collateral_value} 万元" if profile.has_collateral else "无",
         "纳税记录", "有" if profile.has_tax_record else "无"),
        ("开票流水", "有" if profile.has_invoice else "无",
         "贷款用途", PURPOSE_CN.get(profile.loan_purpose.value, "-")),
        ("贷款需求", f"{profile.loan_amount} 万元", "期望期限", f"{profile.preferred_term_months} 个月"),
        ("是否急用", "是" if profile.urgent else "否", "", ""),
    ]
    for ri, (k1, v1, k2, v2) in enumerate(rows, 1):
        ws2.cell(row=ri, column=1, value=k1).fill = LABEL_FILL
        ws2.cell(row=ri, column=1).font = LABEL_FONT
        ws2.cell(row=ri, column=2, value=v1)
        ws2.cell(row=ri, column=3, value=k2).fill = LABEL_FILL
        ws2.cell(row=ri, column=3).font = LABEL_FONT
        ws2.cell(row=ri, column=4, value=v2)
        for c in range(1, 5):
            ws2.cell(row=ri, column=c).border = BORDER
    _set_widths(ws2, [16, 28, 16, 28])

    # ---------- 推荐方案 ----------
    ws3 = wb.create_sheet("推荐方案")
    head = ["排名", "产品", "机构类型", "预估额度(万)", "年化利率", "建议期限(月)",
            "预估月供(万)", "预估总利息(万)", "通过率", "匹配分", "放款时效", "是否需抵押"]
    ws3.append(head)
    _style_header(ws3, 1, len(head))
    for i, p in enumerate(result.plans, 1):
        ws3.append([
            i, p.product_name, p.provider_type, p.estimated_amount,
            f"{p.annual_rate_min}%-{p.annual_rate_max}%", p.suggested_term_months,
            p.monthly_payment_estimate, p.total_interest_estimate,
            p.approval_probability, p.score, p.expected_release_days,
            "需要" if p.requires_collateral else "免抵押",
        ])
        if i == 1:
            for c in range(1, len(head) + 1):
                ws3.cell(row=ws3.max_row, column=c).fill = BEST_FILL
    for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, max_col=len(head)):
        for cell in row:
            cell.border = BORDER
    _set_widths(ws3, [6, 22, 14, 14, 16, 14, 14, 16, 10, 9, 18, 12])

    # 最优方案推荐理由
    if result.plans:
        start = ws3.max_row + 2
        ws3.cell(row=start, column=1, value=f"最优推荐【{result.plans[0].product_name}】推荐理由").font = LABEL_FONT
        for j, reason in enumerate(result.plans[0].match_reasons, 1):
            ws3.cell(row=start + j, column=1, value=f"· {reason}")
            ws3.merge_cells(start_row=start + j, start_column=1, end_row=start + j, end_column=6)

    # ---------- 风控评估 ----------
    ws4 = wb.create_sheet("风控评估")
    ws4.append(["综合风控评分", f"{r.score} / 100", "风险等级", f"{r.grade}（{r.grade_label}）"])
    if r.debt_ratio is not None:
        ws4.append(["负债杠杆", f"{int(r.debt_ratio * 100)}%", "", ""])
    ws4.append([])
    hr = ws4.max_row + 1
    ws4.append(["风险因子", "影响", "说明"])
    _style_header(ws4, hr, 3)
    for f in r.factors:
        ws4.append([f.name, IMPACT_CN.get(f.impact, f.impact), f.detail])
        ws4.cell(row=ws4.max_row, column=3).alignment = WRAP
    for row in ws4.iter_rows(min_row=hr, max_row=ws4.max_row, max_col=3):
        for cell in row:
            cell.border = BORDER
    _set_widths(ws4, [20, 12, 70])

    # ---------- 个性化建议 ----------
    ws5 = wb.create_sheet("个性化建议")
    ws5["A1"] = "个性化融资建议"
    ws5["A1"].font = TITLE_FONT
    for j, a in enumerate(result.personalized_advice, 1):
        cell = ws5.cell(row=j + 2, column=1, value=f"{j}. {a}")
        cell.alignment = WRAP
        ws5.merge_cells(start_row=j + 2, start_column=1, end_row=j + 2, end_column=4)
    _set_widths(ws5, [30, 30, 30, 30])

    # ---------- 补贴政策 ----------
    if result.subsidies:
        ws6 = wb.create_sheet("扶持政策")
        head = ["政策名称", "类别", "主管部门", "扶持内容", "申请要点"]
        ws6.append(head)
        _style_header(ws6, 1, len(head))
        for s in result.subsidies:
            ws6.append([s.name, s.category, s.authority, s.benefit, s.apply_points])
            ws6.cell(row=ws6.max_row, column=4).alignment = WRAP
            ws6.cell(row=ws6.max_row, column=5).alignment = WRAP
        for row in ws6.iter_rows(min_row=1, max_row=ws6.max_row, max_col=len(head)):
            for cell in row:
                cell.border = BORDER
        _set_widths(ws6, [24, 14, 22, 44, 44])

    # ---------- 资质提升建议 ----------
    if result.improvement_tips:
        ws7 = wb.create_sheet("提升建议")
        ws7["A1"] = "资质提升建议"
        ws7["A1"].font = TITLE_FONT
        for j, tip in enumerate(result.improvement_tips, 1):
            cell = ws7.cell(row=j + 2, column=1, value=f"{j}. {tip}")
            cell.alignment = WRAP
            ws7.merge_cells(start_row=j + 2, start_column=1, end_row=j + 2, end_column=4)
        _set_widths(ws7, [30, 30, 30, 30])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
