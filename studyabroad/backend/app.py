"""出国读研选校匹配 - FastAPI 后端。

提供院校/专业/项目数据、匹配、账号体系、收藏清单、申请时间线与文书清单，
并托管前端静态页面。独立于同仓库的贷款助手项目。
"""
import hashlib
import io
import json
import os
import re
import secrets
import sqlite3
from contextlib import closing
from datetime import datetime
from typing import List, Optional

from fastapi import FastAPI, File, Header, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "studyabroad.db")

app = FastAPI(title="出国读研选校匹配", version="2.0.0")
app.add_middleware(
    CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"]
)


def load_json(name: str) -> dict:
    with open(os.path.join(DATA_DIR, name), encoding="utf-8") as f:
        return json.load(f)


def db():
    return closing(sqlite3.connect(DB_PATH))


def init_db() -> None:
    with db() as conn:
        conn.execute(
            """CREATE TABLE IF NOT EXISTS users (
                username TEXT PRIMARY KEY,
                pwd_hash TEXT NOT NULL,
                salt TEXT NOT NULL,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )"""
        )
        conn.execute(
            """CREATE TABLE IF NOT EXISTS tokens (
                token TEXT PRIMARY KEY,
                username TEXT NOT NULL,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )"""
        )
        conn.execute(
            """CREATE TABLE IF NOT EXISTS shortlist (
                username TEXT NOT NULL,
                program_id TEXT NOT NULL,
                status TEXT DEFAULT '待定',
                deadline TEXT DEFAULT '',
                note TEXT DEFAULT '',
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (username, program_id)
            )"""
        )
        # 旧库平滑升级：补列
        cols = {r[1] for r in conn.execute("PRAGMA table_info(shortlist)").fetchall()}
        for col, ddl in [
            ("status", "ALTER TABLE shortlist ADD COLUMN status TEXT DEFAULT '待定'"),
            ("deadline", "ALTER TABLE shortlist ADD COLUMN deadline TEXT DEFAULT ''"),
            ("note", "ALTER TABLE shortlist ADD COLUMN note TEXT DEFAULT ''"),
        ]:
            if col not in cols:
                conn.execute(ddl)
        conn.commit()


init_db()

TIER_DEFAULT = "双非"
TIER_ORDER = ["保底", "匹配", "冲刺", "超出"]


# ===================== 账号体系 =====================
def hash_pwd(password: str, salt: str) -> str:
    return hashlib.pbkdf2_hmac(
        "sha256", password.encode(), salt.encode(), 100_000
    ).hex()


def resolve_user(token: Optional[str]) -> str:
    """token -> username；游客（无 token）返回 'guest'。token 非法则 401。"""
    if not token:
        return "guest"
    with db() as conn:
        row = conn.execute(
            "SELECT username FROM tokens WHERE token=?", (token,)
        ).fetchone()
    if not row:
        raise HTTPException(status_code=401, detail="登录已失效，请重新登录")
    return row[0]


class AuthIn(BaseModel):
    username: str
    password: str


@app.post("/api/register")
def api_register(body: AuthIn):
    username = body.username.strip()
    if len(username) < 2 or len(body.password) < 4:
        raise HTTPException(status_code=400, detail="用户名至少2位、密码至少4位")
    salt = secrets.token_hex(8)
    pwd_hash = hash_pwd(body.password, salt)
    with db() as conn:
        exists = conn.execute(
            "SELECT 1 FROM users WHERE username=?", (username,)
        ).fetchone()
        if exists:
            raise HTTPException(status_code=409, detail="用户名已存在")
        conn.execute(
            "INSERT INTO users (username, pwd_hash, salt) VALUES (?,?,?)",
            (username, pwd_hash, salt),
        )
        token = secrets.token_hex(16)
        conn.execute(
            "INSERT INTO tokens (token, username) VALUES (?,?)", (token, username)
        )
        conn.commit()
    return {"token": token, "username": username}


@app.post("/api/login")
def api_login(body: AuthIn):
    username = body.username.strip()
    with db() as conn:
        row = conn.execute(
            "SELECT pwd_hash, salt FROM users WHERE username=?", (username,)
        ).fetchone()
        if not row or hash_pwd(body.password, row[1]) != row[0]:
            raise HTTPException(status_code=401, detail="用户名或密码错误")
        token = secrets.token_hex(16)
        conn.execute(
            "INSERT INTO tokens (token, username) VALUES (?,?)", (token, username)
        )
        conn.commit()
    return {"token": token, "username": username}


@app.post("/api/logout")
def api_logout(authorization: Optional[str] = Header(None)):
    token = (authorization or "").replace("Bearer ", "").strip()
    if token:
        with db() as conn:
            conn.execute("DELETE FROM tokens WHERE token=?", (token,))
            conn.commit()
    return {"ok": True}


# ===================== 软实力背景评分 =====================
TOP_EMPLOYERS = [
    "腾讯", "阿里", "字节", "抖音", "华为", "百度", "美团", "京东", "网易", "小米",
    "蚂蚁", "快手", "拼多多", "携程", "滴滴", "商汤", "旷视", "大疆",
    "google", "microsoft", "amazon", "apple", "meta", "facebook", "ibm", "intel", "nvidia",
    "麦肯锡", "贝恩", "波士顿", "bcg", "普华永道", "德勤", "安永", "毕马威", "四大",
    "高盛", "摩根", "中金", "中信", "招商", "goldman", "morgan",
]


def is_top_employer(name: Optional[str]) -> bool:
    s = (name or "").lower()
    return any(e.lower() in s for e in TOP_EMPLOYERS)


def compute_soft(soft: Optional[dict]) -> dict:
    """计算软背景竞争力加成，与前端 match.js computeSoftBackground 保持一致。"""
    if not soft:
        return {"score": 0, "boost": 0.0, "level": "未填写", "highlights": []}
    boost = 0.0
    hl: List[str] = []

    # 实习：兼容结构化条目 internshipList 或数量 internships
    intern_list = [
        x for x in (soft.get("internshipList") or [])
        if x and (x.get("company") or x.get("role") or x.get("desc"))
    ]
    intern_count = len(intern_list) or (soft.get("internships") or 0)
    intern_top = soft.get("internshipTop") or any(
        is_top_employer(x.get("company")) for x in intern_list
    )
    if intern_count > 0:
        v = min(intern_count, 3) * 1
        if intern_top:
            v += 0.8
        boost += v
        hl.append(f"{intern_count} 段实习" + ("（含名企/大厂）" if intern_top else ""))

    paper_w = {"SCI/EI": 4, "核心": 2.5, "会议": 1.5, "在投": 0.8}.get(
        soft.get("paperLevel"), 0
    )
    if (soft.get("papers") or 0) > 0 and paper_w:
        boost += min(soft["papers"], 3) * paper_w
        hl.append(f"{soft['papers']} 篇{soft['paperLevel']}论文")
    comp_w = {"国际/国家级": 2.5, "省级": 1.2, "校级": 0.4}.get(
        soft.get("competitionLevel"), 0
    )
    if (soft.get("competitions") or 0) > 0 and comp_w:
        boost += min(soft["competitions"], 3) * comp_w
        hl.append(f"{soft['competitions']} 项{soft['competitionLevel']}竞赛获奖")
    if (soft.get("research") or 0) > 0:
        boost += min(soft["research"], 2) * 1
        hl.append(f"{soft['research']} 段科研经历")
    if (soft.get("workYears") or 0) > 0:
        boost += min(soft["workYears"], 4) * 0.5
        hl.append(f"{soft['workYears']} 年工作经验")

    # 项目 / 校园活动 / 交换经历
    def count_filled(arr):
        return len([
            x for x in (arr or [])
            if x and any(str(v or "").strip() for v in x.values())
        ])
    proj_n = count_filled(soft.get("projectList"))
    if proj_n > 0:
        boost += min(proj_n, 3) * 0.6
        hl.append(f"{proj_n} 个项目经历")
    act_n = count_filled(soft.get("activityList"))
    if act_n > 0:
        boost += min(act_n, 2) * 0.4
        hl.append(f"{act_n} 项校园活动")
    ex_n = count_filled(soft.get("exchangeList"))
    if ex_n > 0:
        boost += min(ex_n, 2) * 0.7
        hl.append(f"{ex_n} 段交换/海外经历")

    # 证书
    certs = [c.strip() for c in re.split(r"[\n;,，；]", soft.get("certificates") or "") if c.strip()]
    if certs:
        boost += min(len(certs), 3) * 0.5
        hl.append(f"{len(certs)} 项证书")

    boost = min(boost, 6)
    score = round(boost / 6 * 100)
    level = "强" if score >= 70 else "中等" if score >= 35 else "一般" if score > 0 else "未填写"
    return {"score": score, "boost": round(boost, 1), "level": level, "highlights": hl}


# ===================== 匹配算法 =====================
def evaluate(program: dict, profile: dict) -> dict:
    req = program["requirements"]
    tier = profile.get("tier") or TIER_DEFAULT
    avg_by_tier = req.get("avgByTier", {})
    required_avg = avg_by_tier.get(tier, avg_by_tier.get(TIER_DEFAULT, 80))
    soft = profile.get("_soft") or compute_soft(profile.get("soft"))
    avg_gap = (profile.get("avg") or 0) - required_avg
    eff_gap = avg_gap + soft["boost"]

    ielts_req = req.get("ielts") or {"overall": 6.5, "sub": 6.0}
    user_overall = (profile.get("ielts") or {}).get("overall", 0) or 0
    user_sub = (profile.get("ielts") or {}).get("sub", 0) or user_overall
    ielts_provided = user_overall > 0
    ielts_ok = (not ielts_provided) or (
        user_overall >= ielts_req["overall"] and user_sub >= ielts_req["sub"]
    )

    gre_req_raw = req.get("gre")
    # gre 可能是 {"total":...} 结构化要求，也可能是自由文本（如 "GRE 建议"）。
    # 仅当为结构化数值要求时才做硬性比较，文本型仅作提示。
    gre_req = gre_req_raw if isinstance(gre_req_raw, dict) else None
    gre_note = gre_req_raw if isinstance(gre_req_raw, str) else None
    user_gre = (profile.get("gre") or {}).get("total", 0) or 0
    gre_provided = user_gre > 0
    gre_ok = (not gre_req) or (not gre_provided) or user_gre >= gre_req["total"]

    if eff_gap >= 3:
        category = "保底"
    elif eff_gap >= 0:
        category = "匹配"
    elif eff_gap >= -3:
        category = "冲刺"
    else:
        category = "超出"

    warnings: List[str] = []
    coop = profile.get("coop") or None
    lang_waivable = bool(coop and coop.get("englishTaught"))
    if ielts_provided and not ielts_ok:
        warnings.append(
            f"雅思未达标：需总分 {ielts_req['overall']}/小分 {ielts_req['sub']}，"
            f"你当前 {user_overall}/{user_sub}"
            + ("（你的全英文授课海外学位通常可申请豁免，如适用则不受影响）" if lang_waivable else "")
        )
        if not lang_waivable:
            if category == "保底":
                category = "匹配"
            elif category == "匹配":
                category = "冲刺"
    if gre_req and not gre_provided:
        warnings.append(f"该项目通常需要 GRE（建议总分 {gre_req['total']}+），你未填写")
    if gre_req and gre_provided and not gre_ok:
        warnings.append(f"GRE 偏低：建议总分 {gre_req['total']}+，你当前 {user_gre}")
        if category == "保底":
            category = "匹配"
        elif category == "匹配":
            category = "冲刺"
    if gre_note and not gre_provided:
        warnings.append(f"GRE：{gre_note}（未填写，建议确认目标项目要求）")

    coop_notes: List[str] = []
    if coop:
        coop_notes.append(
            f"🎓 本科为中外合作大学，申研以合作方【{coop.get('partner')}（{coop.get('degreeRegion')}）】的海外学位申请"
        )
        if coop.get("englishTaught"):
            if not ielts_provided:
                coop_notes.append(
                    f"全英文授课海外学位，多数 {coop.get('degreeRegion')}/英港新院校可申请雅思/托福豁免，请向目标院校确认"
                )
            else:
                coop_notes.append("全英文授课海外学位，多数院校可申请语言豁免，请向目标院校确认是否仍需提交雅思")
        gs = coop.get("gradeSystem")
        if gs == "UK":
            coop_notes.append(
                "成绩按英国荣誉学位评估：一等(First)≥70、二等一(2:1)60–69、二等二(2:2)50–59；英港新院校常按学位等级（而非国内百分制）评估"
            )
        elif gs == "US":
            coop_notes.append("成绩按美国 4.0 GPA 体系评估，部分项目仍要求 GRE/GMAT，请按 4.0 制填写均分")
        elif gs == "HK":
            coop_notes.append("成绩按香港院校体系（GPA/荣誉等级）评估，全英文授课背景申港/英/新常可豁免语言")

    return {
        "program": program,
        "category": category,
        "requiredAvg": required_avg,
        "avgGap": round(avg_gap, 1),
        "softBoost": soft["boost"],
        "softLevel": soft["level"],
        "ieltsReq": ielts_req,
        "ieltsOk": ielts_ok,
        "greReq": gre_req,
        "warnings": warnings,
        "coopNotes": coop_notes,
    }


def match(programs: List[dict], profile: dict) -> dict:
    buckets = {k: [] for k in TIER_ORDER}
    profile["_soft"] = compute_soft(profile.get("soft"))
    countries = profile.get("countries") or []
    only_verified = profile.get("onlyVerified")

    # 研究生方向过滤：支持多选（取并集）。仅保留库内真实存在的方向，
    # 自定义/未知方向不参与硬过滤（视为不限制），避免误杀。
    valid_fields = {p["field"] for p in programs}
    wanted = [f for f in (profile.get("fields") or []) if f and f != "全部"]
    single = profile.get("field")
    if not wanted and single and single != "全部":
        wanted = [single]
    field_filter = [f for f in wanted if f in valid_fields]

    for program in programs:
        if countries and program["country"] not in countries:
            continue
        if field_filter and program["field"] not in field_filter:
            continue
        if only_verified and not (program.get("provenance") or {}).get("verified"):
            continue
        res = evaluate(program, profile)
        buckets[res["category"]].append(res)

    def rank(r):
        q = r["program"].get("qsRank")
        return q if q else 9999

    buckets["保底"].sort(key=rank)
    buckets["匹配"].sort(key=rank)
    buckets["冲刺"].sort(key=lambda r: -r["avgGap"])
    buckets["超出"].sort(key=lambda r: -r["avgGap"])
    return buckets


# ===================== 数据 / 匹配接口 =====================
class IeltsModel(BaseModel):
    overall: float = 0
    sub: float = 0


class GreModel(BaseModel):
    total: int = 0


class InternshipModel(BaseModel):
    company: str = ""
    role: str = ""
    period: str = ""
    desc: str = ""


class ExperienceModel(BaseModel):
    """通用经历条目（项目/活动/交换）：字段名灵活，允许额外键。"""
    class Config:
        extra = "allow"
    period: str = ""
    desc: str = ""


class SoftModel(BaseModel):
    internshipList: List[InternshipModel] = []
    projectList: List[ExperienceModel] = []
    activityList: List[ExperienceModel] = []
    exchangeList: List[ExperienceModel] = []
    internships: int = 0
    internshipTop: bool = False
    papers: int = 0
    paperLevel: Optional[str] = None
    competitions: int = 0
    competitionLevel: Optional[str] = None
    research: int = 0
    workYears: float = 0
    certificates: str = ""


class CoopModel(BaseModel):
    class Config:
        extra = "allow"
    school: str = ""
    partner: str = ""
    partnerEn: str = ""
    degreeRegion: str = ""
    gradeSystem: str = ""
    englishTaught: bool = False
    note: str = ""


class Profile(BaseModel):
    tier: str = TIER_DEFAULT
    avg: float = 0
    field: str = "全部"
    fields: List[str] = []
    countries: List[str] = []
    onlyVerified: bool = False
    coop: Optional[CoopModel] = None
    ielts: IeltsModel = IeltsModel()
    gre: GreModel = GreModel()
    soft: SoftModel = SoftModel()


@app.get("/api/universities")
def api_universities():
    return load_json("universities.json")


@app.get("/api/majors")
def api_majors():
    return load_json("majors.json")


# ===================== 本科专业 → 研究生方向智能识别 =====================
# 规则化「AI 识别」：用关键词把任意本科专业名映射到相关研究生方向，
# 即便该专业不在 majors.json 预设列表中也能给出推荐。
GRAD_FIELD_KEYWORDS = [
    (["人工智能", "机器学习", "深度学习", "数据科学", "大数据", "数据挖掘",
      "artificial intelligence", "machine learning", "data science", "big data", " ai "],
     ["数据科学/AI", "计算机/CS", "商业分析"]),
    (["智能科学", "智能技术", "智能系统", "智能", "intelligence"],
     ["数据科学/AI", "计算机/CS", "电子/电气工程"]),
    (["软件", "software"], ["软件工程", "计算机/CS", "信息系统/IT"]),
    (["物联网", "网络工程", "信息技术", "信息工程", "information technology", "iot"],
     ["信息系统/IT", "计算机/CS", "通信工程"]),
    (["计算机", "计算机科学", "程序", "编程", "computer", "computing"],
     ["计算机/CS", "软件工程", "数据科学/AI", "信息系统/IT"]),
    (["统计", "应用数学", "数学", "statistic", "mathematic"],
     ["数据科学/AI", "商业分析", "金融工程", "经济学"]),
    (["金融工程", "金融数学", "量化", "financial engineering", "quantitative"],
     ["金融工程", "金融", "数据科学/AI"]),
    (["金融", "投资", "银行", "证券", "保险", "finance", "investment", "banking"],
     ["金融", "金融工程", "经济学", "商业分析"]),
    (["会计", "审计", "财务", "account", "audit"],
     ["会计", "金融", "商业分析"]),
    (["经济", "贸易", "economic", "trade"],
     ["经济学", "金融", "商业分析", "公共政策/管理"]),
    (["市场", "营销", "广告", "品牌", "marketing", "advertis", "brand"],
     ["市场营销", "管理学/商科", "传媒"]),
    (["工商", "管理", "人力资源", "物流", "供应链", "电子商务", "旅游", "酒店",
      "business", "management", "logistic", "hospitality", "mba", "hr"],
     ["管理学/商科", "商业分析", "市场营销", "公共政策/管理"]),
    (["通信", "信号", "无线", "communication", "telecom", "signal"],
     ["通信工程", "电子/电气工程", "计算机/CS"]),
    (["电子工程", "电子信息", "电子科学", "电气", "集成电路", "微电子", "自动化",
      "electric", "electronic"],
     ["电子/电气工程", "通信工程", "计算机/CS"]),
    (["机械", "车辆", "汽车", "动力", "能源", "制造", "机器人", "航空", "航天",
      "mechanical", "automotive", "manufactur", "robot", "aerospace"],
     ["机械工程", "材料工程"]),
    (["材料", "高分子", "冶金", "纳米", "复合材料", "material", "metallurg", "polymer"],
     ["材料工程", "机械工程"]),
    (["土木", "建筑", "结构", "城乡规划", "测绘", "岩土", "交通工程",
      "civil", "architect", "structural"],
     ["土木/建筑", "设计/艺术", "环境科学"]),
    (["环境", "生态", "可持续", "气候", "地理", "资源", "海洋",
      "environment", "ecolog", "sustainab", "climate", "geograph"],
     ["环境科学", "公共政策/管理", "土木/建筑"]),
    (["新闻", "传播", "媒体", "广播", "影视", "出版", "media", "journalism", "communication studies"],
     ["传媒", "市场营销", "翻译/语言"]),
    (["教育", "师范", "教学", "学前", "教育学", "educat", "teaching", "pedagog"],
     ["教育", "心理学", "公共政策/管理"]),
    (["法学", "法律", "法", "知识产权", " law", "legal", "juris"],
     ["法律(LLM)", "公共政策/管理"]),
    (["心理", "psycholog"], ["心理学", "教育", "数据科学/AI"]),
    (["公共", "行政", "政治", "国际关系", "政府", "社会工作", "社会学",
      "public", "politic", "administration", "international relation", "sociolog"],
     ["公共政策/管理", "管理学/商科", "经济学"]),
    (["设计", "艺术", "美术", "视觉", "工业设计", "数字媒体", "动画", "服装",
      "design", "art", "visual", "animation"],
     ["设计/艺术", "传媒", "计算机/CS"]),
    (["英语", "翻译", "语言", "文学", "汉语", "外语", "日语", "法语", "德语",
      "linguistic", "translat", "language", "literature", "tesol"],
     ["翻译/语言", "教育", "传媒"]),
    (["生物", "医", "药", "护理", "食品", "化学", "化工", "物理", "生命科学",
      "biolog", "medic", "pharma", "nursing", "chemi", "physics"],
     ["数据科学/AI", "材料工程", "环境科学"]),
]
# 兜底通用推荐（无任何关键词命中时）
FALLBACK_GRAD_FIELDS = ["管理学/商科", "商业分析", "数据科学/AI", "公共政策/管理"]


def recommend_fields_for_major(major: str) -> dict:
    """根据本科专业名（任意输入）识别相关研究生方向。"""
    name = (major or "").strip()
    if not name:
        return {"major": name, "matched": "empty", "fields": [], "note": ""}

    # 1) 预设列表精确/包含匹配
    majors = load_json("majors.json")["undergradMajors"]
    flat = name.lower()
    exact = next((m for m in majors if m["name"] == name), None)
    if exact:
        return {"major": name, "matched": "exact",
                "fields": exact.get("recommendFields", []),
                "note": ""}
    contained = next(
        (m for m in majors if m["name"] in name or name in m["name"]), None
    )
    if contained:
        return {"major": name, "matched": "similar",
                "fields": contained.get("recommendFields", []),
                "note": f"按相近专业「{contained['name']}」推荐"}

    # 2) 关键词智能识别
    picked: List[str] = []
    for keywords, fields in GRAD_FIELD_KEYWORDS:
        if any(k.strip().lower() in flat for k in keywords):
            for f in fields:
                if f not in picked:
                    picked.append(f)
    if picked:
        return {"major": name, "matched": "ai",
                "fields": picked[:6],
                "note": "AI 智能识别（该专业不在预设列表，以下为相关方向推荐）"}

    # 3) 兜底
    return {"major": name, "matched": "fallback",
            "fields": FALLBACK_GRAD_FIELDS,
            "note": "未能精确识别该专业，以下为通用方向推荐，可自行输入更精确的方向"}


class RecommendIn(BaseModel):
    major: str = ""


@app.post("/api/recommend-fields")
def api_recommend_fields(body: RecommendIn):
    return recommend_fields_for_major(body.major)


@app.get("/api/programs")
def api_programs():
    return load_json("programs.json")


@app.post("/api/match")
def api_match(profile: Profile):
    data = load_json("programs.json")
    return match(data["programs"], profile.model_dump())


# ===================== 收藏清单（按用户隔离） =====================
class ShortlistItem(BaseModel):
    program_id: str


@app.get("/api/shortlist")
def api_get_shortlist(authorization: Optional[str] = Header(None)):
    user = resolve_user((authorization or "").replace("Bearer ", "").strip())
    programs = {p["id"]: p for p in load_json("programs.json")["programs"]}
    out = []
    with db() as conn:
        rows = conn.execute(
            "SELECT program_id, status, deadline, note FROM shortlist "
            "WHERE username=? ORDER BY created_at DESC",
            (user,),
        ).fetchall()
    for pid, status, deadline, note in rows:
        if pid not in programs:
            continue
        p = dict(programs[pid])
        p["track"] = {"status": status or "待定", "deadline": deadline or "", "note": note or ""}
        out.append(p)
    return out


@app.post("/api/shortlist")
def api_add_shortlist(item: ShortlistItem, authorization: Optional[str] = Header(None)):
    user = resolve_user((authorization or "").replace("Bearer ", "").strip())
    with db() as conn:
        conn.execute(
            "INSERT OR IGNORE INTO shortlist (username, program_id) VALUES (?, ?)",
            (user, item.program_id),
        )
        conn.commit()
    return {"ok": True}


class TrackUpdate(BaseModel):
    program_id: str
    status: Optional[str] = None
    deadline: Optional[str] = None
    note: Optional[str] = None


VALID_STATUS = {"待定", "准备中", "已递交", "面试中", "已录取", "已拒绝"}


@app.patch("/api/shortlist")
def api_update_track(body: TrackUpdate, authorization: Optional[str] = Header(None)):
    user = resolve_user((authorization or "").replace("Bearer ", "").strip())
    if body.status is not None and body.status not in VALID_STATUS:
        raise HTTPException(status_code=400, detail="无效的申请状态")
    sets, vals = [], []
    if body.status is not None:
        sets.append("status=?"); vals.append(body.status)
    if body.deadline is not None:
        sets.append("deadline=?"); vals.append(body.deadline)
    if body.note is not None:
        sets.append("note=?"); vals.append(body.note)
    if not sets:
        return {"ok": True}
    vals += [user, body.program_id]
    with db() as conn:
        conn.execute(
            f"UPDATE shortlist SET {', '.join(sets)} WHERE username=? AND program_id=?",
            vals,
        )
        conn.commit()
    return {"ok": True}


@app.delete("/api/shortlist")
def api_del_shortlist(
    program_id: str, authorization: Optional[str] = Header(None)
):
    user = resolve_user((authorization or "").replace("Bearer ", "").strip())
    with db() as conn:
        conn.execute(
            "DELETE FROM shortlist WHERE username=? AND program_id=?",
            (user, program_id),
        )
        conn.commit()
    return {"ok": True}


# ===================== 申请时间线 + 文书清单 =====================
def build_checklist(programs: List[dict]) -> dict:
    docs = [
        {"name": "成绩单（中英文，学校盖章）", "required": True},
        {"name": "在读证明 / 毕业证 + 学位证（中英文）", "required": True},
        {"name": "个人简历 CV", "required": True},
        {"name": "个人陈述 PS / Statement of Purpose", "required": True},
        {"name": "推荐信 2–3 封（学术 / 实习导师）", "required": True},
        {"name": "雅思 / 托福 语言成绩单", "required": True},
        {"name": "护照（首页）", "required": True},
    ]
    need_gre = any(p["requirements"].get("gre") for p in programs)
    need_portfolio = any(
        p["field"] in ("设计/艺术", "土木/建筑") or "作品集" in (p["requirements"].get("notes") or "")
        for p in programs
    )
    need_test = any(p["field"] == "翻译/语言" for p in programs)
    if need_gre:
        docs.append({"name": "GRE / GMAT 成绩（部分美国/商科项目要求）", "required": False})
    if need_portfolio:
        docs.append({"name": "作品集 Portfolio（设计/建筑类）", "required": False})
    if need_test:
        docs.append({"name": "翻译 / 口译入学测试（语言类项目）", "required": False})

    countries = sorted(set(p["country"] for p in programs))
    timeline = [
        {"phase": "前一年 3–6 月", "task": "确定选校清单、准备并报考雅思/托福（及 GRE）"},
        {"phase": "前一年 7–8 月", "task": "撰写 CV / PS、联系推荐人、准备成绩单等材料"},
        {"phase": "前一年 9–10 月", "task": "网申陆续开放，英国/香港/新加坡建议第一轮递交"},
        {"phase": "前一年 11–12 月", "task": "美国早申/截止，递交主批次，部分项目面试"},
        {"phase": "当年 1–3 月", "task": "陆续放榜，补充材料 / 参加面试"},
        {"phase": "当年 4–6 月", "task": "确定 offer、缴占位费、办理签证与住宿"},
        {"phase": "当年 7–9 月", "task": "行前准备、体检、机票，入学报到"},
    ]
    country_notes = []
    tips = {
        "英国": "滚动录取，越早递交越占优，热门项目可能提前满位",
        "美国": "多在 12–1 月截止，需提前规划 GRE 与文书，部分有早申",
        "中国香港": "分轮录取，建议首轮，名额有限",
        "新加坡": "名额少、竞争大，11 月起尽早递交",
        "澳大利亚": "全年多轮、可滚动申请，但建议提前半年锁定",
    }
    for c in countries:
        if c in tips:
            country_notes.append({"country": c, "tip": tips[c]})

    return {
        "documents": docs,
        "timeline": timeline,
        "countryNotes": country_notes,
        "programCount": len(programs),
    }


@app.get("/api/checklist")
def api_checklist(authorization: Optional[str] = Header(None)):
    user = resolve_user((authorization or "").replace("Bearer ", "").strip())
    programs = {p["id"]: p for p in load_json("programs.json")["programs"]}
    with db() as conn:
        rows = conn.execute(
            "SELECT program_id FROM shortlist WHERE username=?", (user,)
        ).fetchall()
    chosen = [programs[r[0]] for r in rows if r[0] in programs]
    return build_checklist(chosen)


class ChecklistReq(BaseModel):
    program_ids: List[str] = []


@app.post("/api/checklist")
def api_checklist_post(body: ChecklistReq):
    """游客模式：直接传入项目 id 列表生成清单（不依赖账号）。"""
    programs = {p["id"]: p for p in load_json("programs.json")["programs"]}
    chosen = [programs[i] for i in body.program_ids if i in programs]
    return build_checklist(chosen)


# ===================== 简历 PDF 智能解析 =====================
PAPER_TOP = ["SCI", "EI", "SSCI", "Nature", "Science", "IEEE", "ACL", "CVPR", "NeurIPS", "ICML"]
COMP_NATIONAL = ["国家级", "国际", "全国", "ACM", "ICPC", "挑战杯", "数学建模", "美赛", "国奖", "国家奖学金", "Kaggle"]
COMP_PROVINCE = ["省级", "省赛", "省一", "省二", "省三"]
CERT_KEYWORDS = [
    "CFA", "CPA", "FRM", "ACCA", "CMA", "PMP", "教师资格", "法律职业资格", "司法考试",
    "计算机等级", "计算机二级", "计算机三级", "软考", "CET-4", "CET-6", "四级", "六级",
    "托业", "TOEIC", "BEC", "驾驶证", "普通话", "证券从业", "基金从业", "银行从业",
]


def extract_resume_text(raw: bytes) -> str:
    try:
        from pypdf import PdfReader
    except Exception:  # pragma: no cover
        raise HTTPException(status_code=500, detail="服务器缺少 PDF 解析组件 pypdf")
    try:
        reader = PdfReader(io.BytesIO(raw))
        return "\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"PDF 解析失败：{e}")


def parse_resume(text: str) -> dict:
    """规则化 NLP：从简历纯文本中抽取结构化背景。"""
    out: dict = {"raw_chars": len(text)}
    flat = re.sub(r"\s+", "", text)
    low = text.lower()

    unis = load_json("universities.json")["universities"]
    school = next((u for u in unis if u["name"] in flat), None)
    if not school:  # 去掉后缀再试（如“武汉大学”→“武汉”过宽，故仅整名匹配）
        for u in unis:
            short = u["name"].replace("大学", "").replace("学院", "")
            if len(short) >= 3 and short in flat:
                school = u
                break
    if school:
        out["school"] = school["name"]
        out["tier"] = school["tier"]
        out["province"] = school["province"]

    majors = load_json("majors.json")["undergradMajors"]
    major = next((m for m in majors if m["name"] in flat), None)
    if major:
        out["undergradMajor"] = major["name"]
        out["recommendFields"] = major.get("recommendFields", [])

    m = re.search(r"(?:GPA|绩点|平均绩点)\D{0,4}([0-4]\.\d{1,2})", text, re.I)
    if m:
        out["gpa"] = float(m.group(1))
    m = re.search(r"(?:均分|平均分|加权|百分制|WAM)\D{0,4}(\d{2,3}(?:\.\d{1,2})?)", text, re.I)
    if m:
        v = float(m.group(1))
        if 50 <= v <= 100:
            out["avg"] = v

    m = re.search(r"(?:雅思|IELTS)\D{0,4}([4-9](?:\.\d)?)", text, re.I)
    if m:
        out["ielts"] = float(m.group(1))
    m = re.search(r"(?:托福|TOEFL)\D{0,4}(\d{2,3})", text, re.I)
    if m:
        out["toefl"] = int(m.group(1))
    m = re.search(r"GRE\D{0,4}(\d{3})", text, re.I)
    if m and 260 <= int(m.group(1)) <= 340:
        out["gre"] = int(m.group(1))

    # 实习：尝试从文本行中抽取「公司 + 岗位」结构化条目
    intern_list = []
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    role_kw = r"(实习生|实习|intern|工程师|开发|分析师|助理|顾问|经理|专员|研究员|developer|engineer|analyst|consultant|assistant)"
    for ln in lines:
        if not re.search(r"实习|intern", ln, re.I):
            continue
        company = next((e for e in TOP_EMPLOYERS if e.lower() in ln.lower()), "")
        if not company:
            mco = re.search(r"([\u4e00-\u9fa5A-Za-z][\u4e00-\u9fa5A-Za-z&·]{1,20}?(?:公司|集团|科技|银行|证券|研究院|事务所|实验室|Inc|Ltd|Corp))", ln)
            if mco:
                company = mco.group(1)
        mrole = re.search(role_kw, ln, re.I)
        role = mrole.group(0) if mrole else ""
        mperiod = re.search(
            r"20\d{2}[.\-/年]\d{0,2}\s*[-–~至]+\s*(?:20\d{2}[.\-/年]?\d{0,2}|至今|present|now)",
            ln, re.I,
        )
        period = mperiod.group(0) if mperiod else ""
        # 跳过纯小标题行（如「实习经历：」），要求有公司，或岗位不是泛指的“实习”
        if not company and role in ("", "实习", "intern", "实习经历"):
            continue
        if company or role:
            intern_list.append({"company": company, "role": role, "period": period, "desc": ""})
    # 去重（按公司+岗位）
    seen = set()
    uniq = []
    for it in intern_list:
        key = (it["company"], it["role"])
        if key not in seen:
            seen.add(key)
            uniq.append(it)
    intern_hits = len(re.findall(r"实习|intern", text, re.I))
    if uniq:
        out["internshipList"] = uniq[:5]
        out["internships"] = len(uniq[:5])
        out["internshipTop"] = any(is_top_employer(x["company"]) for x in uniq)
    elif intern_hits:
        out["internships"] = min(intern_hits, 5)
        out["internshipTop"] = is_top_employer(low)

    # 论文
    paper_kw = len(re.findall(r"论文|发表|paper|published|期刊|conference|会议", text, re.I))
    if paper_kw:
        out["papers"] = min(max(1, len(re.findall(r"论文|paper|published", text, re.I))), 5)
        if any(k.lower() in low for k in [p.lower() for p in PAPER_TOP]):
            out["paperLevel"] = "SCI/EI"
        elif re.search(r"核心|CSSCI|CPCI|北大核心", text, re.I):
            out["paperLevel"] = "核心"
        elif re.search(r"在投|under\s*review|投稿", text, re.I):
            out["paperLevel"] = "在投"
        else:
            out["paperLevel"] = "会议"

    # 竞赛
    comp_kw = len(re.findall(r"竞赛|大赛|获奖|奖项|award|competition|奖学金", text, re.I))
    if comp_kw:
        out["competitions"] = min(comp_kw, 5)
        if any(k.lower() in low for k in [c.lower() for c in COMP_NATIONAL]):
            out["competitionLevel"] = "国际/国家级"
        elif any(k in text for k in COMP_PROVINCE):
            out["competitionLevel"] = "省级"
        else:
            out["competitionLevel"] = "校级"

    # 科研
    research_hits = len(re.findall(r"科研|课题|实验室|research|RA\b|项目研究", text, re.I))
    if research_hits:
        out["research"] = min(research_hits, 4)

    # 证书 / 资格
    certs = []
    for kw in CERT_KEYWORDS:
        if kw.lower() in low and kw not in certs:
            certs.append(kw)
    if certs:
        out["certificates"] = "\n".join(certs)

    fields = [
        k for k in ["school", "undergradMajor", "avg", "gpa", "ielts", "toefl", "gre",
                    "internships", "papers", "competitions", "research", "certificates"]
        if k in out
    ]
    out["extracted"] = fields
    return out


@app.post("/api/extract-resume")
async def api_extract_resume(file: UploadFile = File(...)):
    if not (file.filename or "").lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="请上传 PDF 格式的简历")
    raw = await file.read()
    if len(raw) > 8 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="文件过大，请控制在 8MB 以内")
    text = extract_resume_text(raw)
    if len(text.strip()) < 20:
        raise HTTPException(
            status_code=422,
            detail="未能从 PDF 提取到文字（可能是扫描件/图片版简历），请改用文本版 PDF 或手动填写",
        )
    return parse_resume(text)


# ===================== AI 文书辅助（PS 提纲 + CV 建议） =====================
def build_sop(programs: List[dict], profile: dict) -> dict:
    soft = compute_soft(profile.get("soft"))
    fields = sorted(set(p["field"] for p in programs)) or [profile.get("field") or "目标方向"]
    countries = sorted(set(p["country"] for p in programs))
    target_names = [p["university"] for p in programs][:6]
    primary_field = fields[0]

    ps_outline = [
        {
            "section": "1. 开篇动机（Hook & Motivation）",
            "tips": [
                f"用一个具体经历切入你为何选择「{primary_field}」方向，避免空泛的“我从小热爱”。",
                "点明你想解决的真实问题或行业痛点，建立个人叙事主线。",
            ],
        },
        {
            "section": "2. 学术基础（Academic Background）",
            "tips": [
                f"结合本科专业「{profile.get('undergradMajor') or '相关专业'}」说明核心课程与成绩亮点。"
                + (f"（你的均分约 {profile.get('avg')} 分可作为佐证）" if profile.get("avg") else ""),
                "挑 1–2 门与目标方向最相关的课程/项目深入讲，而非罗列成绩单。",
            ],
        },
        {
            "section": "3. 实践与科研（Experience）",
            "tips": [],
        },
        {
            "section": "4. 为何选择该项目（Why This Program）",
            "tips": [
                f"针对 {('、'.join(target_names) or '目标院校')} 分别提到具体课程、导师或研究中心，做到“一校一版”。",
                "把项目特色与你的职业目标对应起来，体现匹配度（fit）。",
            ],
        },
        {
            "section": "5. 职业规划（Career Goal）",
            "tips": [
                "给出清晰的短期（毕业 1–3 年）与长期目标，并说明该硕士项目如何成为跳板。",
                f"结合 {('、'.join(countries) or '目标地区')} 的行业环境，让规划更可信。",
            ],
        },
    ]

    exp_tips = []
    s0 = profile.get("soft") or {}
    intern_n = len(s0.get("internshipList") or []) or s0.get("internships") or 0
    if soft["highlights"]:
        exp_tips.append("用 STAR 法则（情境-任务-行动-结果）量化你的：" + "、".join(soft["highlights"]) + "。")
    if s0.get("papers"):
        exp_tips.append("论文部分写清你承担的具体工作（数据/实验/写作）与成果，而非只列标题。")
    if intern_n:
        names = "、".join(
            [x.get("company") for x in (s0.get("internshipList") or []) if x.get("company")][:3]
        )
        exp_tips.append(
            (f"实习（{names}）" if names else "实习经历")
            + "突出可量化结果（如提升 X%、处理 Y 量级数据）与方法论迁移能力。"
        )
    if (s0.get("certificates") or "").strip():
        exp_tips.append("相关证书/资格可佐证专业能力，在 CV 单列且在文书中点到为止（说明与目标方向的关联）。")
    if (s0.get("projectList") or []):
        pnames = "、".join([x.get("name") for x in s0["projectList"] if x.get("name")][:3])
        exp_tips.append(
            (f"项目（{pnames}）" if pnames else "项目经历")
            + "讲清你的角色、技术难点与可量化产出，体现解决问题的能力。"
        )
    if (s0.get("exchangeList") or []):
        exp_tips.append("交换 / 海外经历可佐证跨文化适应力与英语环境学习能力，简要点明学术收获。")
    if (s0.get("activityList") or []):
        exp_tips.append("校园活动 / 社团经历用于展现领导力与团队协作，挑 1 项最有分量的写，避免流水账。")
    if not exp_tips:
        exp_tips.append("软背景较弱：建议补充 1 段相关实习或科研/课程项目，并在文书中重点呈现可迁移能力。")
    ps_outline[2]["tips"] = exp_tips

    cv_bullets = []
    s = profile.get("soft") or {}
    if s.get("papers"):
        cv_bullets.append(f"Published {s['papers']} paper(s) [{s.get('paperLevel','')}] — 注明你的署名位次与贡献。")
    if s.get("competitions"):
        cv_bullets.append(f"Won {s['competitions']} award(s) [{s.get('competitionLevel','')}] — 写明奖项级别与名次/比例。")
    intern_list = s.get("internshipList") or []
    if intern_list:
        for it in intern_list[:3]:
            head = " / ".join([x for x in [it.get("company"), it.get("role"), it.get("period")] if x]) or "实习经历"
            cv_bullets.append(f"{head} — 用 2–3 条量化 bullet 描述职责与成果。")
    elif s.get("internships"):
        cv_bullets.append(f"{s['internships']} internship(s)" + ("（含名企）" if s.get("internshipTop") else "") + " — 每段 2–3 条量化 bullet。")
    if s.get("research"):
        cv_bullets.append(f"{s['research']} research experience(s) — 写明研究问题、方法、你的角色与产出。")
    for it in (s.get("projectList") or [])[:3]:
        head = " / ".join([x for x in [it.get("name"), it.get("role"), it.get("period")] if x]) or "项目经历"
        cv_bullets.append(f"{head} — 注明技术栈与可量化成果。")
    for it in (s.get("exchangeList") or [])[:2]:
        head = " / ".join([x for x in [it.get("school"), it.get("program"), it.get("period")] if x]) or "交换经历"
        cv_bullets.append(f"{head} — 写明修读课程 / 学术收获。")
    acts = s.get("activityList") or []
    if acts:
        head = " / ".join([x for x in [acts[0].get("org"), acts[0].get("role")] if x]) or "校园活动"
        cv_bullets.append(f"{head} 等 {len(acts)} 项活动 — 突出领导力与组织成果。")
    certs = [c.strip() for c in re.split(r"[\n;,，；]", s.get("certificates") or "") if c.strip()]
    if certs:
        cv_bullets.append("Certifications：" + "、".join(certs) + " — 单列于 CV「证书」栏。")
    cv_bullets.append("统一动词开头 + 量化结果，控制在 1 页 A4，时间倒序。")

    competitiveness = {
        "level": soft["level"],
        "score": soft["score"],
        "boost": soft["boost"],
        "comment": {
            "强": "软背景突出，建议适当冲刺更高排名项目，文书重点放在科研/成果深度。",
            "中等": "软背景良好，保持冲刺+匹配组合，文书强化最亮的 1–2 项经历。",
            "一般": "软背景偏基础，建议尽快补充 1 段实习/科研，文书突出潜力与匹配度。",
            "未填写": "尚未填写软背景，完善后可获得更精准的竞争力评估与文书建议。",
        }[soft["level"]],
    }

    return {
        "psOutline": ps_outline,
        "cvBullets": cv_bullets,
        "competitiveness": competitiveness,
        "programCount": len(programs),
        "fields": fields,
        "disclaimer": "本辅助为基于你的背景与选校的结构化写作指引，不代写文书内容，请结合自身真实经历撰写。",
    }


class SopReq(BaseModel):
    program_ids: List[str] = []
    profile: Optional[dict] = None


@app.post("/api/sop")
def api_sop(body: SopReq, authorization: Optional[str] = Header(None)):
    programs_all = {p["id"]: p for p in load_json("programs.json")["programs"]}
    ids = body.program_ids
    if not ids:
        user = resolve_user((authorization or "").replace("Bearer ", "").strip())
        with db() as conn:
            rows = conn.execute(
                "SELECT program_id FROM shortlist WHERE username=?", (user,)
            ).fetchall()
        ids = [r[0] for r in rows]
    chosen = [programs_all[i] for i in ids if i in programs_all]
    return build_sop(chosen, body.profile or {})


# ===================== 申请规划导出（PDF / Excel） =====================
class ExportItem(BaseModel):
    program_id: str
    status: str = "待定"
    deadline: str = ""
    note: str = ""


class ExportReq(BaseModel):
    items: List[ExportItem] = []


def _collect_export(items: List[ExportItem]) -> List[dict]:
    programs = {p["id"]: p for p in load_json("programs.json")["programs"]}
    out = []
    for it in items:
        if it.program_id not in programs:
            continue
        p = dict(programs[it.program_id])
        p["track"] = {"status": it.status or "待定", "deadline": it.deadline or "", "note": it.note or ""}
        out.append(p)
    return out


@app.post("/api/export/shortlist.xlsx")
def export_xlsx(body: ExportReq):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    items = _collect_export(body.items)
    wb = Workbook()
    ws = wb.active
    ws.title = "申请清单"
    headers = ["国家/地区", "院校", "项目", "方向", "QS", "申请状态", "截止日期",
               "参考均分门槛(双非)", "雅思要求", "学费", "时长", "备注"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="111111")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
    for p in items:
        req = p["requirements"]
        ielts = req.get("ielts") or {}
        ws.append([
            p["country"], p["university"], p["program"], p["field"],
            p.get("qsRank") or "-", p["track"]["status"], p["track"]["deadline"] or "-",
            (req.get("avgByTier") or {}).get("双非", "-"),
            f"{ielts.get('overall','-')}/{ielts.get('sub','-')}" if ielts else "见官网",
            p.get("tuition", "-"), p.get("duration", "-"), p["track"]["note"] or "",
        ])
    widths = [10, 22, 30, 14, 6, 10, 12, 16, 12, 18, 8, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"studyabroad_shortlist_{datetime.now():%Y%m%d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.post("/api/export/plan.pdf")
def export_pdf(body: ExportReq):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )

    items = _collect_export(body.items)
    try:
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        cn_font = "STSong-Light"
    except Exception:
        cn_font = "Helvetica"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm,
                            leftMargin=14 * mm, rightMargin=14 * mm)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("t", parent=styles["Title"], fontName=cn_font, fontSize=18)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontName=cn_font, fontSize=12)
    body = ParagraphStyle("b", parent=styles["Normal"], fontName=cn_font, fontSize=9, leading=13)

    story = [Paragraph("出国读研 · 申请规划清单", title),
             Paragraph(f"生成时间：{datetime.now():%Y-%m-%d %H:%M} · 共 {len(items)} 个项目", body),
             Spacer(1, 8)]

    data = [["院校", "项目", "方向", "状态", "截止", "雅思", "学费"]]
    for p in items:
        req = p["requirements"]
        ielts = req.get("ielts") or {}
        data.append([
            Paragraph(p["university"], body),
            Paragraph(p["program"], body),
            Paragraph(p["field"], body),
            Paragraph(p["track"]["status"], body),
            Paragraph(p["track"]["deadline"] or "-", body),
            Paragraph(f"{ielts.get('overall','-')}/{ielts.get('sub','-')}" if ielts else "见官网", body),
            Paragraph(p.get("tuition", "-"), body),
        ])
    table = Table(data, colWidths=[34 * mm, 38 * mm, 20 * mm, 16 * mm, 18 * mm, 16 * mm, 30 * mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), cn_font),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f4f5")]),
    ]))
    story.append(table)
    story.append(Spacer(1, 10))

    chk = build_checklist(items)
    story.append(Paragraph("文书 / 材料清单", h2))
    for d in chk["documents"]:
        mark = "□" if not d["required"] else "■"
        story.append(Paragraph(f"{mark} {d['name']}" + ("" if d["required"] else "（按需）"), body))
    story.append(Spacer(1, 8))
    story.append(Paragraph("申请时间线", h2))
    for t in chk["timeline"]:
        story.append(Paragraph(f"· {t['phase']}：{t['task']}", body))

    story.append(Spacer(1, 10))
    story.append(Paragraph(load_json("programs.json")["meta"]["disclaimer"], body))

    doc.build(story)
    buf.seek(0)
    fname = f"studyabroad_plan_{datetime.now():%Y%m%d}.pdf"
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ===================== 静态页面托管 =====================
# 仅暴露前端所需的 css / js / data 目录；backend（含数据库）与 scripts 不对外暴露
app.mount("/css", StaticFiles(directory=os.path.join(BASE_DIR, "css")), name="css")
app.mount("/js", StaticFiles(directory=os.path.join(BASE_DIR, "js")), name="js")
app.mount("/data", StaticFiles(directory=os.path.join(BASE_DIR, "data")), name="data")


@app.get("/")
def index():
    return FileResponse(os.path.join(BASE_DIR, "index.html"))
